import cv2
import glob
import pyzbar
import pandas as pd
from openpyxl import Workbook

# Function to read barcodes from an image and return a DataFrame with the data


def read_barcodes(image):
    # Scan the image for barcodes
    barcodes = pyzbar.decode(image)

    # Create a pandas DataFrame with the barcode data
    df = pd.DataFrame([barcode.data.decode("utf-8")
                      for barcode in barcodes], columns=["Barcode Data"])

    return df


# Get a list of image files in the "barcodes" directory
image_files = glob.glob("barcodes/*.png")

# Create an Excel file and write the DataFrames to it
book = Workbook()

# Read barcodes from each image and write to separate sheet
for i, image_file in enumerate(image_files):
    # Read the image file
    image = cv2.imread(image_file)

    # Create a sheet for the image
    sheet = book.create_sheet("Barcodes {}".format(i + 1))

    # Read barcodes from image and write to sheet
    df = read_barcodes(image)
    sheet.append(df.columns)
    for row in df.values:
        sheet.append(row)

book.save("barcodes.xlsx")

# Function to read barcodes from an image and return a DataFrame with the data


def read_barcodes(image):
    # Scan the image for barcodes
    barcodes = pyzbar.decode(image)

    # Create a pandas DataFrame with the barcode data
    df = pd.DataFrame([barcode.data.decode("utf-8")
                      for barcode in barcodes], columns=["Barcode Data"])

    return df


# Get a list of image files in the "barcodes" directory
image_files = glob.glob("barcodes/*.png")

# Create an Excel file and write the DataFrames to it
book = Workbook()

# Read barcodes from each image and write to separate sheet
for i, image_file in enumerate(image_files):
    # Read the image file
    image = cv2.imread(image_file)

    # Create a sheet for the image
    sheet = book.create_sheet("Barcodes {}".format(i + 1))

    # Read barcodes from image and write to sheet
    df = read_barcodes(image)
    sheet.append(df.columns)
    for row in df.values:
        sheet.append(row)

book.save("barcodes.xlsx")
